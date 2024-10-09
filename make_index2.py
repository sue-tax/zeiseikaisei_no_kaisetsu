"""
税制改正の解説の目次用のExcelを作成する

https://github.com/juu7g/Python-PDF2text
pdf_PDF2text
を参考にしている
"""

from pdfminer.high_level import extract_pages
from pdfminer.layout import LAParams, LTTextBox
import collections
import os, sys, argparse, glob

import re
import openpyxl


class MakeIndex():
    """
    税制改正の解説の目次用のExcelを作成する。
    """

    def __init__(self, argv:list):
        """
        コンストラクタ

        Args:
            argv:   以下
                    ベースフォルダ名
        """
        self.footer = 60    # フッターのy座標。ページの最下部が0。これより下の位置の文字は抽出しない
        self.header = 1000  # ヘッダーのy座標。 これより上の位置の文字は抽出しない

        if not argv:
            return

        # コマンドライン引数の解析
        parser = argparse.ArgumentParser()		# インスタンス作成
        parser.add_argument('base_folder', type=str, help="ベースフォルダ名")	# 引数定義
                
        args = parser.parse_args(argv)				# 引数の解析
        # print(args)						# 引数の参照
        self.base_folder = args.base_folder
        
        self.p_dir = re.compile('([A-Z][0-9]{2})_([0-9]{4})')
        self.p_pdf = re.compile('([0-9]{4})-([0-9]{4})')
        
        '''
        令和六年能登半島地震災害の被災者に係る所得税法及び災害被害
        本　文
        者に対する租税の減免、徴収猶予等に関する法律の臨時特例に関
        参照頁
        する法律の創設
        　に対応できず
        '''
        self.p_filename = re.compile(r'.+?[pP]([0-9]+)[-_][pP]?([0-9]+)[^.]*\.pdf')

        self.p_low = re.compile(r'((本　文 ?|詳[ 　]解 ?)(\r\n|\n|\r))?' \
                                '([^\r\n]+?)' \
                                '(\r\n|\n|\r)?' \
                                '(本　文 ?|詳[ 　]解 ?)?' \
                                '(\r\n|\n|\r)' \
                                '(参照頁|ページ) ?(\r\n|\n|\r)')
        # self.p_low = re.compile(r'本　文 ?(\r\n|\n|\r)' \
        #                         '(.+?)' \
        #                         '(\r\n|\n|\r)' \
        #                         '参照頁 ?(\r\n|\n|\r)')
        self.p_dai = re.compile(r'(第[一二三四五六七八九]) ?　([^\r\n]+?)(\r\n|\n|\r)')

        # self.p_suji = re.compile(r'([0-9０-９]+) ?　([^\r\n]+?)(\r\n|\n|\r)')
        # TODO 2 ％,3 月31 3 年12月31日 3 分の1以上の弊害あるが、2006対応
        self.p_suji_2007e = re.compile(r'([0-9０-９]+) ?　([^\r\n]+?)(\r\n|\n|\r)')
        self.p_suji_2006 = re.compile(r'([0-9０-９]+) ?[ 　]([^\r\n]+?)(\r\n|\n|\r)')
        self.pos_suji = 2
        # self.p_suji = re.compile(r'([0-9０-９]+) ?[ 　](?!(％|月31日))([^\r\n]+?)(\r\n|\n|\r)')
        # self.pos_suji = 3
        
        self.p_kakko = re.compile(r'([⑴-⒇]) ?[ 　]([^\r\n]+?)(\r\n|\n|\r)')
        self.p_maru = re.compile(r'([①-⑳]) ?[ 　]([^\r\n]+?)(\r\n|\n|\r)')
        
        # self.p_page = re.compile(r'([0-9]+)(\r\n|\n|\r)')
        # 実際にページ数が全角になることがあった
        self.p_page = re.compile(r'([0-9０-９]+) ?(\r\n|\n|\r)')
        self.p_through = re.compile(r'[^\r\n]+?(\r\n|\n|\r)')
        
        self.trans_zenhan = str.maketrans('１２３４５６７８９０',
                '1234567890')
        # self.trans_2006_kakko = str.maketrans('盧盪蘯盻眈眇眄眩眤眞眥眛眷眸睇睚睨睫睛睥睿',
        self.trans_2006_kakko = str.maketrans('盧盪蘯盻眈眇眄眩眤眞眥眛眷眸睇睚睨睫睛睥',
                '⑴⑵⑶⑷⑸⑹⑺⑻⑼⑽⑾⑿⒀⒁⒂⒃⒄⒅⒆⒇')
        self.trans_2007_kakko = str.maketrans('asdfghjklmnopqrtuvwx',
                '⑴⑵⑶⑷⑸⑹⑺⑻⑼⑽⑾⑿⒀⒁⒂⒃⒄⒅⒆⒇')


    def create_file_list(self, files, m_dir):
        self.dict_file = {}
        # print(files)
        if int(self.str_seireki) != 2005:
            for file in files:
                # print(file)
                
                if int(self.str_seireki) == 2006:
                    num_min = 61
                    num_max = 726
                else:
                    m_filename = self.p_filename.match(file)
                    if m_filename == None:
                        # print(file)
                        continue
                    # print(m_filename.group(1))
                    # print(m_filename.group(2))
                    num_min = int(m_filename.group(1))
                    num_max = int(m_filename.group(2))
                # print(num_min, num_max)
                for i in range(num_min, num_max+1):
                    file_name = os.path.basename(file)
                    str_link = r'https://www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/PDF/' + file_name
                    if int(self.str_seireki) == 2006:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/9551815/www.mof.go.jp/tax_policy/tax_reform/outline/fy2006/' \
                            + file_name
                    if int(self.str_seireki) <= 2015:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/9551815/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/pdf/' + file_name
                    if 2010 <= int(self.str_seireki) <= 2011:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/9551815/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/PDF/' + file_name
                    if 2016 <= int(self.str_seireki) <= 2019:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/11344177/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/pdf/' + file_name
                    if 2020 <= int(self.str_seireki) <= 2020:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/11551246/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/pdf/' + file_name
                    if 2021 <= int(self.str_seireki) <= 2021:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/11719722/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/' + file_name
                    str_file = './' + self.str_wareki+'_' + \
                            self.str_seireki + '/' + file_name
                    self.dict_file[i] = (file, num_min, num_max,
                            str_link, str_file, file_name)
        else:
            # H17 2005
            for file, num_min, num_max in zip(files,
                    [ 49, 64, 92, 111, 137, 160, 179, 205, 243, 286, 321, 338, 353, 371, 383, 390, 430, 462],
                    [ 63, 91, 110, 136, 159, 178, 204, 242, 285, 320, 337, 352, 370, 382, 389, 429, 461, 480]):
                for i in range(num_min, num_max+1):
                    file_name = os.path.basename(file)
                    str_link = r'https://warp.ndl.go.jp/info:ndljp/pid/1022127/www.mof.go.jp/finance/f1708betu/' + \
                            file_name
                    str_file = './' + self.str_wareki+'_' + \
                            self.str_seireki + '/' + file_name
                    self.dict_file[i] = (file, num_min, num_max,
                            str_link, str_file, file_name)
        
    def make_pdf_to_index(self):
        if not os.path.exists(self.base_folder):
            print("フォルダ {} がありません".self.base_folder)
            return
        
        dirs = glob.glob(self.base_folder + "/*/")
        # print(dirs)
        
        wb = openpyxl.Workbook()
        self.ws = wb.active
        self.ws.append(('和暦', '西暦', '税法等',
                '第', '項目',
                '数字', '項目',
                'カッコ', '項目',
                '丸数字', '項目',
                '頁', 'リンク'
                ))
        self.ws_row = 2
        
        # md
        
        for str_dir in dirs:
            # print(str_dir)
            list_dir = str_dir.split('\\')
            # m_dir = self.p_dir.match(list_dir[-2])
            m_dir = self.p_dir.fullmatch(list_dir[-2])
            if m_dir == None:
                continue
            # print(m_dir)
            self.str_wareki = m_dir.group(1)
            self.str_seireki = m_dir.group(2)

            if int(self.str_seireki) != 2005:
                # files = glob.glob(str_dir + 'p*-*.pdf')
                # pxxxx-pyyyy.pdf, Pxxx-Pyyy.pdf
                files = glob.glob(str_dir + '*-*.pdf')
                files2 = glob.glob(str_dir + '*_*.pdf')
                files.extend(files2)
                # print(files)
                files.sort()
                # print(files)
    
                self.create_file_list(files, m_dir)
                            
                out_text = self.base_folder + '/' + list_dir[-2] + '.txt'
                # print(out_text)
                if int(self.str_seireki) != 2006:
                    self.make_each_pdf_to_text(files[1], out_text)
                else:
                    # self.make_each_pdf_to_text("p011-p061.pdf", out_text)
                    self.make_each_pdf_to_text(files[0], out_text)
                # out_excel =  self.base_folder + '/' + list_dir[-2] + '.xlsx'
                # out_md =  self.base_folder + '/' + list_dir[-2] + '.xlsx'
                
                self.make_each_text_to_index(out_text)
            else:
                # print(str_dir)
                files = [ str_dir + "2.pdf", str_dir + "3.pdf",
                         str_dir + "4.pdf", str_dir + "5.pdf", str_dir + "6.pdf",
                         str_dir + "7.pdf", str_dir + "8.pdf", str_dir + "9.pdf",
                         str_dir + "10.pdf", str_dir + "11.pdf",
                         str_dir + "12.pdf", str_dir + "13.pdf",
                         str_dir + "14.pdf", str_dir + "15.pdf", str_dir + "16.pdf",
                         str_dir + "17.pdf", str_dir + "18.pdf", str_dir + "19.pdf" ]
                self.create_file_list(files, m_dir)
                out_text = self.base_folder + '/' + list_dir[-2] + '.txt'
                self.make_each_pdf_to_text(str_dir + "1-2.pdf", out_text)
                self.make_each_text_to_index(out_text)
                # exit(99)

        str_filter = "A1:K" + str(self.ws_row - 1)
        self.ws.auto_filter.ref = str_filter
        self.ws.freeze_panes = 'D2'
        self.ws.column_dimensions['A'].width = 4.5
        self.ws.column_dimensions['B'].width = 4.5
        self.ws.column_dimensions['C'].width = 40
        self.ws.column_dimensions['D'].width = 5
        self.ws.column_dimensions['E'].width = 25
        self.ws.column_dimensions['F'].width = 3.25
        self.ws.column_dimensions['G'].width = 65
        self.ws.column_dimensions['H'].width = 3.25
        self.ws.column_dimensions['I'].width = 65
        self.ws.column_dimensions['J'].width = 3.25
        self.ws.column_dimensions['K'].width = 65
        self.ws.column_dimensions['L'].width = 4.5
        self.ws.column_dimensions['M'].width = 25
        
        out_excel =  self.base_folder + '/' + '税制改正の解説' + '.xlsx'
        # out_md =  self.base_folder + '/' + '税制改正の解説' + '.xlsx'

        wb.save(out_excel)

     
    def make_each_text_to_index(self, in_text):
        f_text = open(in_text, "r", encoding="utf_8")
        str_text = f_text.read()
        f_text.close()
        # print(in_text)
        
        if int(self.str_seireki) == 2006:
            self.p_suji = self.p_suji_2006
        else:
            self.p_suji = self.p_suji_2007e
        for m_low in self.p_low.finditer(str_text):
            # print(m_low)
            # num_low = m_low.group(1)
            # print(m_low.groups())
            str_low = m_low.group(4)
            if int(self.str_seireki) == 2005:
                str_low = str_low.replace(' ', '')
            # print(num_low)
            # print(str_low)
            
            before_page = ''
            num_page = ''
            
            offset = m_low.end()

            num_dai = str_dai = num_suji = str_suji = \
                    num_kakko = str_kakko = num_maru = str_maru = ''

            m_dai = self.p_dai.match(str_text, offset)
            while True:
                if m_dai:
                    offset = m_dai.end()
                    num_dai = m_dai.group(1)
                    str_dai = m_dai.group(2)
                    # print(num_dai)
                    # print(str_dai)
                else:
                    num_dai = ''
                    str_dai = ''
            
                num_suji = str_suji = num_kakko = str_kakko = num_maru = str_maru = ''
                m_suji = self.p_suji.match(str_text, offset)
                while not m_suji:
                    m_page = self.p_page.match(str_text, offset)
                    # print("0", m_page)
                    if m_page == None:
                        num_page = before_page
                    else:
                        # print(m_page.group(1))
                        offset = m_page.end()
                        num_page = m_page.group(1).translate(self.trans_zenhan)
                        m_suji = self.p_suji.match(str_text, offset)
                        before_page = num_page
                        continue
                    m_dai = self.p_dai.match(str_text, offset)
                    m_low = self.p_low.match(str_text, offset)
                    if not(m_dai or m_low):
                        m_through = self.p_through.match(str_text, offset)
                        if m_through:
                            offset = m_through.end()
                            before_page = ''
                            m_suji = self.p_suji.match(str_text, offset)
                            continue
                        break
                    else:
                        break
                while m_suji:
                    # print("A", m_suji)
                    # print("A", num_page, before_page, m_suji)
                    offset = m_suji.end()
                    num_suji = m_suji.group(1).translate(self.trans_zenhan)
                    # print(m_suji.groups())
                    str_suji = m_suji.group(self.pos_suji)
                    # before_page = ''
                    m_page = self.p_page.match(str_text, offset)
                    if m_page == None:
                        num_page = before_page
                        if num_page != '':
                            data_tuple = (self.str_wareki, self.str_seireki,
                                    str_low, num_dai, str_dai,
                                    num_suji, str_suji,
                                    '', '',
                                    '', '',
                                    num_page)
                            self.write_excel(data_tuple)
                            # print("☆", data_tuple)
                        before_page = ''
                    else:
                        # print(m_page.group(1))
                        offset = m_page.end()
                        num_page = m_page.group(1).translate(self.trans_zenhan)
                        data_tuple = (self.str_wareki, self.str_seireki,
                                str_low, num_dai, str_dai,
                                num_suji, str_suji,
                                '', '',
                                '', '',
                                num_page)
                        # print(data_tuple)
                        self.write_excel(data_tuple)
                        # print("☆", data_tuple)
                        m_suji = self.p_suji.match(str_text, offset)
                        before_page = num_page
                    m_kakko = self.p_kakko.match(str_text, offset)
                    m_maru = self.p_maru.match(str_text, offset)
                    while (not m_kakko) and (not m_maru):
                        m_page = self.p_page.match(str_text, offset)
                        if m_page == None:
                            num_page = before_page
                        else:
                            # print(m_page.group(1))
                            offset = m_page.end()
                            num_page = m_page.group(1).translate(self.trans_zenhan)
                            m_kakko = self.p_kakko.match(str_text, offset)
                            m_maru = self.p_maru.match(str_text, offset)
                            before_page = num_page
                            continue
                        m_suji = self.p_suji.match(str_text, offset)
                        m_dai = self.p_dai.match(str_text, offset)
                        m_low = self.p_low.match(str_text, offset)
                        # カッコ数字がなく、丸数字のパターンあり
                        m_maru = self.p_maru.match(str_text, offset)
                        if not(m_maru or m_suji or m_dai or m_low):
                            m_through = self.p_through.match(str_text, offset)
                            if m_through:
                                offset = m_through.end()
                                before_page = ''
                                m_kakko = self.p_kakko.match(str_text, offset)
                                m_maru = self.p_maru.match(str_text, offset)
                                continue
                            break
                        else:
                            break
                    if not m_maru: 
                        while m_kakko:
                            # print("B0", m_kakko)
                            offset = m_kakko.end()
                            num_kakko = m_kakko.group(1)
                            str_kakko = m_kakko.group(2)
                            # print("before_page", before_page)
                            m_page = self.p_page.match(str_text, offset)
                            if m_page == None:
                                num_page = before_page
                                if num_page != '':
                                    data_tuple = (self.str_wareki, self.str_seireki,
                                            str_low, num_dai, str_dai,
                                            num_suji, str_suji,
                                            num_kakko, str_kakko,
                                            '', '',
                                            num_page)
                                    self.write_excel(data_tuple)
                                    # print("☆", data_tuple)
                                before_page = ''
                            else:
                                # print(m_page.group(1))
                                offset = m_page.end()
                                num_page = m_page.group(1).translate(self.trans_zenhan)
                                data_tuple = (self.str_wareki, self.str_seireki,
                                        str_low, num_dai, str_dai,
                                        num_suji, str_suji,
                                        num_kakko, str_kakko,
                                        '', '',
                                        num_page)
                                self.write_excel(data_tuple)
                                # print("☆", data_tuple)
                                m_kakko = self.p_kakko.match(str_text, offset)
                                before_page = num_page
                            m_maru = self.p_maru.match(str_text, offset)
                            while not m_maru:
                                m_page = self.p_page.match(str_text, offset)
                                if m_page == None:
                                    num_page = before_page
                                else:
                                    # print(m_page.group(1))
                                    offset = m_page.end()
                                    num_page = m_page.group(1).translate(self.trans_zenhan)
                                    m_maru = self.p_maru.match(str_text, offset)
                                    # test
                                    before_page = num_page
                                    continue
                                m_kakko = self.p_kakko.match(str_text, offset)
                                m_suji = self.p_suji.match(str_text, offset)
                                m_dai = self.p_dai.match(str_text, offset)
                                m_low = self.p_low.match(str_text, offset)
                                if not(m_kakko or m_suji or m_dai or m_low):
                                    m_through = self.p_through.match(str_text, offset)
                                    if m_through:
                                        offset = m_through.end()
                                        m_maru = self.p_maru.match(str_text, offset)
                                        continue
                                    break
                                else:
                                    break
                            while m_maru:
                                # print("C0", m_maru)
                                offset = m_maru.end()
                                num_maru = m_maru.group(1)
                                str_maru = m_maru.group(2)
        
                                m_page = self.p_page.match(str_text, offset)
                                if m_page == None:
                                    num_page = before_page
                                    if num_page != '':
                                        data_tuple = (self.str_wareki, self.str_seireki,
                                                str_low, num_dai, str_dai,
                                                num_suji, str_suji,
                                                num_kakko, str_kakko,
                                                num_maru, str_maru,
                                                num_page)
                                        self.write_excel(data_tuple)
                                        # print("☆", data_tuple)
                                    before_page = ''
                                else:
                                    # print(m_page.group(1))
                                    offset = m_page.end()
                                    num_page = m_page.group(1).translate(self.trans_zenhan)
                                    data_tuple = (self.str_wareki, self.str_seireki,
                                            str_low, num_dai, str_dai,
                                            num_suji, str_suji,
                                            num_kakko, str_kakko,
                                            num_maru, str_maru,
                                            num_page)
                                    self.write_excel(data_tuple)
                                    # print("☆", data_tuple)
                                    m_maru = self.p_maru.match(str_text, offset)
                                    before_page = num_page
                                m_maru = self.p_maru.match(str_text, offset)
                                m_kakko = m_suji = m_dai = m_low = None
                                m_through = True
                                while not m_maru:
                                    # print("C1", str_text[offset:offset+20])
                                    m_page = self.p_page.match(str_text, offset)
                                    if m_page:
                                        # print(m_page.group(1))
                                        offset = m_page.end()
                                        before_page = m_page.group(1)
                                        m_maru = self.p_maru.match(str_text, offset)
                                        continue
                                    m_kakko = self.p_kakko.match(str_text, offset)
                                    if m_kakko:
                                        break
                                    m_suji = self.p_suji.match(str_text, offset)
                                    if m_suji:
                                        break
                                    m_dai = self.p_dai.match(str_text, offset)
                                    if m_dai:
                                        break
                                    m_low = self.p_low.match(str_text, offset)
                                    if m_low:
                                        break
                                    m_through = self.p_through.match(str_text, offset)
                                    if m_through == None:
                                        break
                                    offset = m_through.end()
                                    before_page = ''
                                    m_maru = self.p_maru.match(str_text, offset)
                                if m_kakko:
                                    break
                                if m_suji:
                                    break
                                if m_dai:
                                    break
                                if m_low:
                                    break
                                if m_through == None:
                                    break
                                # if m_maru:
                                #     before_page = ''
                            m_kakko = self.p_kakko.match(str_text, offset)
                            m_suji = m_dai = m_low = None
                            m_through = True
                            while not m_kakko:
                                # print("B1", str_text[offset:offset+20])
                                m_page = self.p_page.match(str_text, offset)
                                if m_page:
                                    # print(m_page.group(1))
                                    offset = m_page.end()
                                    before_page = m_page.group(1)
                                    m_kakko = self.p_kakko.match(str_text, offset)
                                    continue
                                m_suji = self.p_suji.match(str_text, offset)
                                if m_suji:
                                    break
                                m_dai = self.p_dai.match(str_text, offset)
                                if m_dai:
                                    break
                                m_low = self.p_low.match(str_text, offset)
                                if m_low:
                                    break
                                m_through = self.p_through.match(str_text, offset)
                                if m_through == None:
                                    break
                                offset = m_through.end()
                                before_page = ''
                                m_kakko = self.p_kakko.match(str_text, offset)
                            if m_suji:
                                break
                            if m_dai:
                                break
                            if m_low:
                                break
                            if m_through == None:
                                break
                            # if m_kakko:
                            #     before_page = ''
                    else:
                        # カッコ数字がなく、丸数字
                        num_kakko = ''
                        str_kakko = ''
                        while m_maru:
                            # print("c0", m_maru)
                            offset = m_maru.end()
                            num_maru = m_maru.group(1)
                            str_maru = m_maru.group(2)
    
                            m_page = self.p_page.match(str_text, offset)
                            if m_page == None:
                                num_page = before_page
                                if num_page != '':
                                    data_tuple = (self.str_wareki, self.str_seireki,
                                            str_low, num_dai, str_dai,
                                            num_suji, str_suji,
                                            num_kakko, str_kakko,
                                            num_maru, str_maru,
                                            num_page)
                                    self.write_excel(data_tuple)
                                    # print("☆", data_tuple)
                                before_page = ''
                            else:
                                # print(m_page.group(1))
                                offset = m_page.end()
                                num_page = m_page.group(1).translate(self.trans_zenhan)
                                data_tuple = (self.str_wareki, self.str_seireki,
                                        str_low, num_dai, str_dai,
                                        num_suji, str_suji,
                                        num_kakko, str_kakko,
                                        num_maru, str_maru,
                                        num_page)
                                self.write_excel(data_tuple)
                                # print("☆", data_tuple)
                                m_maru = self.p_maru.match(str_text, offset)
                                before_page = num_page
                            m_maru = self.p_maru.match(str_text, offset)
                            m_kakko = m_suji = m_dai = m_low = None
                            m_through = True
                            while not m_maru:
                                # print("C1", str_text[offset:offset+20])
                                m_page = self.p_page.match(str_text, offset)
                                if m_page:
                                    # print(m_page.group(1))
                                    offset = m_page.end()
                                    before_page = m_page.group(1)
                                    m_maru = self.p_maru.match(str_text, offset)
                                    continue
                                m_kakko = self.p_kakko.match(str_text, offset)
                                if m_kakko:
                                    break
                                m_suji = self.p_suji.match(str_text, offset)
                                if m_suji:
                                    break
                                m_dai = self.p_dai.match(str_text, offset)
                                if m_dai:
                                    break
                                m_low = self.p_low.match(str_text, offset)
                                if m_low:
                                    break
                                m_through = self.p_through.match(str_text, offset)
                                if m_through == None:
                                    break
                                offset = m_through.end()
                                before_page = ''
                                m_maru = self.p_maru.match(str_text, offset)
                            if m_kakko:
                                break
                            if m_suji:
                                break
                            if m_dai:
                                break
                            if m_low:
                                break
                            if m_through == None:
                                break
                    m_suji = self.p_suji.match(str_text, offset)
                    m_dai = m_low = None
                    m_through = True
                    while not m_suji:
                        # print("B2", str_text[offset:offset+20])
                        m_page = self.p_page.match(str_text, offset)
                        if m_page:
                            offset = m_page.end()
                            before_page = m_page.group(1)
                            m_suji = self.p_suji.match(str_text, offset)
                            continue
                        m_dai = self.p_dai.match(str_text, offset)
                        if m_dai:
                            break
                        m_low = self.p_low.match(str_text, offset)
                        if m_low:
                            break
                        m_through = self.p_through.match(str_text, offset)
                        if m_through == None:
                            break
                        offset = m_through.end()
                        before_page = ''
                        m_suji = self.p_suji.match(str_text, offset)
                    if m_dai:
                        break
                    if m_low:
                        break
                    if m_through == None:
                        break
                    m_suji = self.p_suji.match(str_text, offset)
                    # print("B3", str_text[offset:offset+20])
                if m_low:
                    break
                if m_through == None:
                    break
                m_dai = self.p_dai.match(str_text, offset)
    
    def write_excel(self, data_tuple):
        num_page = data_tuple[11]
        # print(data_tuple)
        self.ws.append(data_tuple)
        if int(self.str_seireki) != 2006:
            # print(self.str_seireki)
            # print(num_page)
            int_page = int(num_page) - int(self.dict_file[int(num_page)][1]) + 1
            # print(int_page)
            link_data = self.dict_file[int(num_page)][3] + '#page=' + str(int_page)
            # print(link_data)
            self.ws.cell(column=13, row=self.ws_row).value = \
                    self.dict_file[int(num_page)][4][2:]
            self.ws.cell(column=13, row=self.ws_row).hyperlink = link_data
            # file_data = self.dict_file[int(num_page)][4] + '#page=' + str(int_page)
            # ws.cell(column=14, row=self.ws_row).hyperlink = file_data
            
            # str_foxid = r"/A page=" + str(int_page) + " " + \
            #         self.dict_file[int(num_page)][0]
    
            str_foxid = self.dict_file[int(num_page)][4][2:]
        else:
            int_page = int(num_page) + 1
            link_data = "https://warp.da.ndl.go.jp/info:ndljp/pid/9551815/www.mof.go.jp/tax_policy/tax_reform/outline/fy2006/f1808betu.pdf" \
                    + '#page=' + str(int_page)
            # print(link_data)
            self.ws.cell(column=13, row=self.ws_row).value = \
                    "H18_2006/f1808betu.pdf"
            self.ws.cell(column=13, row=self.ws_row).hyperlink = link_data
            str_foxid = "H18_2006/f1808betu.pdf"
        # foxit reader
        # self.ws.cell(column=14, row=self.ws_row).value = str_foxid
        # self.ws.cell(column=15, row=self.ws_row).value = int_page
        '''
        Sub linkpdfpage()
            Worksheets("Sheet").Activate
            Text = "C:\Program Files (x86)\Foxit Software\Foxit PDF Reader\FoxitPDFReader.exe /A page=" + Mid(Cells(Selection(1).Row, 15).Value, 1) + " " + ActiveWorkbook.Path + "/" + Cells(Selection(1).Row, 14).Value
            Shell (Text)
        End Sub
        '''
        self.ws_row += 1

    def make_each_pdf_to_text(self, in_file, out_file):
        laparams = LAParams()               # パラメータインスタンス
        laparams.boxes_flow = None          # -1.0（水平位置のみが重要）から+1.0（垂直位置のみが重要）default 0.5
        laparams.word_margin = 0.2          # default 0.1
        laparams.char_margin = 2.0          # default 2.0
        laparams.line_margin = 0            # default 0.5

        # print(in_file)
        
        with open(out_file, "w", encoding="utf-8") as f:
            self.text_ = ""
            
            for page_layout in extract_pages(in_file, maxpages=0, laparams=laparams):    # ファイルにwithしている
                for element in sorted(self.flatten_lttext(page_layout, LTTextBox), key=lambda x: (-x.y1, x.x0)):
                    if element.y1 < self.footer:
                        continue  # フッター位置の文字は抽出しない
                    if element.y0 > self.header:
                        continue  # ヘッダー位置の文字は抽出しない
                    _text =element.get_text()
                    # debug
                    # if self.str_seireki == "2009":
                    #     print("y1:{}, y0:{}, x0:{}, x1:{}■{}".
                    #           format(element.y1, element.y0, element.x0, element.x1, _text))
                    # print(_text)
                    if '(cid:' in _text:
                        _text = re.sub(r"\(cid:([0-9]*)\)",
                                lambda m: chr(ord('⑴')+int(m.group(1))-2), _text) 
                    _text = _text.replace(chr(0x18), '？')
                    _text = _text.replace(chr(0x07), '')
                    
                    # 太字のページ数
                    if self.str_seireki == "2009" or self.str_seireki =="2010":
                        if re.fullmatch(r'[0-9]*(\r\n|\n|\r)', _text):
                            _text = _text[int((len(_text)-1)/2):]
                    if self.str_seireki == "2006":
                        _text = _text.translate(self.trans_2006_kakko)
                    if self.str_seireki == "2007":
                        _text = _text.translate(self.trans_2007_kakko)
                        if "i" in _text:
                            _text = _text.replace("i1", "⑾")
                            _text = _text.replace("i2", "⑿")
                            _text = _text.replace("i3", "⒀")
                            _text = _text.replace("i4", "⒁")
                            _text = _text.replace("i5", "⒂")
                            _text = _text.replace("i6", "⒃")
                            _text = _text.replace("i7", "⒄")
                            _text = _text.replace("i8", "⒅")
                            _text = _text.replace("i9", "⒆")
                    self.text_ += _text
                self.write2text(f)

    def flatten(self, l):
        """
        ツリー状になっているイテレータをフラットに返すイテレータ
        """
        for el in l:
            if isinstance(el, collections.abc.Iterable) and not isinstance(el, (str, bytes)):
                yield from self.flatten(el)
            else:
                yield el

    def flatten_lttext(self, l, _type):
        """
        ツリー状になっているイテレータをフラットに返すイテレータ
        返る要素の型を指定
        pdfminerのextract_pagesで使用するのを想定
        要素の型が引数で指定した型を継承したもののみを返す

        Args:
            l:      pdfminerのextract_pages()の戻り値
            _type:  戻したい値の型
        """
        for el in l:
            if isinstance(el, (_type)):
                yield el
            else:
                if isinstance(el, collections.abc.Iterable) and not isinstance(el, (str, bytes)):
                    yield from self.flatten_lttext(el, _type)
                else:
                    continue

    def write2text(self, f):
        """
        Args:
            f:      書き込みファイル
        """
        f.write(self.text_)
        self.text_ = ""

    def convert_pdf_to_text(self):
        """
        PDFファイルをテキストに変換
        PDFは2段に段組みされたものも含む
        """

        laparams = LAParams()               # パラメータインスタンス
        laparams.boxes_flow = None          # -1.0（水平位置のみが重要）から+1.0（垂直位置のみが重要）default 0.5
        laparams.word_margin = 0.2          # default 0.1
        laparams.char_margin = 2.0          # default 2.0
        laparams.line_margin = 0            # default 0.5

        with open(self.output_path, "w", encoding="utf-8") as f:
            self.text_l = ""        # 左側の文字列
            self.text_r = ""        # 右側の文字列
            
            print("Analyzing from {} page to {} page(0:to last)".format(self.start_page, self.last_page))
            
            for page_layout in extract_pages(self.input_path, maxpages=0, laparams=laparams):    # ファイルにwithしている
                if page_layout.pageid < self.start_page: continue                   # 指定開始ページより前は飛ばす
                if self.last_page and self.last_page < page_layout.pageid: break    # 指定終了ページ以降は中断
                if self.border == 0:
                    self.border = int(page_layout.width / 2)
                if page_layout.pageid == self.start_page:
                    print("Check on page #{}".format(page_layout.pageid))
                    print("Page Info width:{}, heght:{}".format(page_layout.width, page_layout.height))
                    print("Calc result border: {}, footer: {}".format(self.border, self.footer))
                for element in sorted(self.flatten_lttext(page_layout, LTTextBox), key=lambda x: (-x.y1, x.x0)):
                    if element.y1 < self.footer: continue  # フッター位置の文字は抽出しない
                    if element.y0 > self.header: continue  # ヘッダー位置の文字は抽出しない
                    _text =element.get_text()
                    # debug
                    # print("y1:{}, y0:{}, x0:{}, x1:{}■{}".
                    #       format(element.y1, element.y0, element.x0, element.x1, _text))

                    if element.x1 < self.border:
                        # 文字列全体が左側
                        self.text_l += _text
                    else:
                        if element.x0 >= self.border:
                            # 文字列全体が右側
                            self.text_r += _text
                        else:
                            # 文字列が境界をまたいでいる場合
                            # 右側に既に文章があれば先に出力する
                            if self.text_r:
                                self.write2text(f)
                            self.text_l += _text
                self.write2text(f)

if __name__ == "__main__":
    make = MakeIndex(sys.argv[1:])
    make.make_pdf_to_index()
