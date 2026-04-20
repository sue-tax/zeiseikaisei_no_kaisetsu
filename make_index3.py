"""
税制改正の解説の目次用のExcelを作成する

https://github.com/juu7g/Python-PDF2text
pdf_PDF2text
を参考にしている
"""

import c
import d
import e

'''
PDFからページ数を正しく読み込めない（コードが\x18）ものは、
9に置き換える。
リンクなどはおかしくなる。
項目内の文章も、変になることがある
'''

from pdfminer.high_level import extract_pages
from pdfminer.layout import LAParams, LTTextBox
import collections
import os, sys, argparse, glob

import re
import openpyxl
from openpyxl.styles.alignment import Alignment


__version__ = 0.11  # 検索用の全項目の列を追加


class MakeIndex():
    """
    税制改正の解説の目次用のExcelを作成する。
    """

    def __init__(self, argv:list):
        """
        コンストラクタ

        Args:
            argv:   以下
                    ベースフォルダ名
        """
        self.footer = 60    # フッターのy座標。ページの最下部が0。これより下の位置の文字は抽出しない
        self.header = 1000  # ヘッダーのy座標。 これより上の位置の文字は抽出しない

        if not argv:
            return

        # コマンドライン引数の解析
        parser = argparse.ArgumentParser()		# インスタンス作成
        parser.add_argument('base_folder', type=str, help="ベースフォルダ名")	# 引数定義
                
        args = parser.parse_args(argv)				# 引数の解析
        self.base_folder = args.base_folder
        
        self.p_dir = re.compile('([A-Z][0-9]{2})_([0-9]{4})')
        self.p_pdf = re.compile('([0-9]{4})-([0-9]{4})')
        
        '''
        令和六年能登半島地震災害の被災者に係る所得税法及び災害被害
        本　文
        者に対する租税の減免、徴収猶予等に関する法律の臨時特例に関
        参照頁
        する法律の創設
        　に対応できず
        '''
        self.p_filename = re.compile(r'.+?[pP]([0-9]+)[-_][pP]?([0-9]+)[^.]*\.pdf')

        self.p_low = re.compile(r'((本　文 ?|詳[ 　]解 ?)(\r\n|\n|\r))?' \
                                '([^\r\n]+?)' \
                                '(\r\n|\n|\r)?' \
                                '(本　文 ?|詳[ 　]解 ?)?' \
                                '(\r\n|\n|\r)' \
                                '(参照頁|ページ) ?(\r\n|\n|\r)')
        self.p_dai = re.compile(r'(第[一二三四五六七八九]) ?　([^\r\n]+?)(\r\n|\n|\r)')

        # ver0.04
        self.p_kansuji = re.compile(r'([一二三四五六七八九]) ?　([^\r\n]+?)(\r\n|\n|\r)')
        
        # TODO 2 ％,3 月31 3 年12月31日 3 分の1以上の弊害あるが、2006対応
        self.p_suji_2007e = re.compile(r'([0-9０-９]+) ?　([^\r\n]+?)(\r\n|\n|\r)')
        self.p_suji_2006 = re.compile(r'([0-9０-９]+) ?[ 　]([^\r\n]+?)(\r\n|\n|\r)')
        self.pos_suji = 2
        
        # self.p_kakko = re.compile(r'[ 　]*([⑴-⒇]) ?[ 　]([^\r\n]+?)(\r\n|\n|\r)')
        self.p_kakko = re.compile(r'[ 　]*([⑴-⒇]|\([0-9]+\)) ?[ 　]([^\r\n]+?)(\r\n|\n|\r)')

        self.p_kakko_mae =  re.compile(r'[ 　]*')
        self.p_kakko_ushiro = re.compile(r' ?[ 　]([^\r\n]+?)(\r\n|\n|\r)')

        self.p_maru = re.compile(r'([ 　]*[①-⑳]) ?[ 　]([^\r\n]+?)(\r\n|\n|\r)')
        
        # 実際にページ数が全角になることがあった
        self.p_page = re.compile(r'([0-9０-９]+) ?(\r\n|\n|\r)')
        self.p_through = re.compile(r'[^\r\n]+?(\r\n|\n|\r)')
        
        self.trans_zenhan = str.maketrans('１２３４５６７８９０',
                '1234567890')

        self.trans_2006_kakko = str.maketrans('盧盪蘯盻眈眇眄眩眤眞眥眛眷眸睇睚睨睫睛睥',
                '⑴⑵⑶⑷⑸⑹⑺⑻⑼⑽⑾⑿⒀⒁⒂⒃⒄⒅⒆⒇')
        self.trans_2007_kakko = str.maketrans('asdfghjklmnopqrtuvwx',
                '⑴⑵⑶⑷⑸⑹⑺⑻⑼⑽⑾⑿⒀⒁⒂⒃⒄⒅⒆⒇')


    def match_kakko(self):
        # (11)～が、別のコードになっている場合の処理
        # d.dprint_method_start()
        m_mae = self.p_kakko_mae.match(self.str_text, self.offset)
        if not m_mae:
            return None
        if m_mae.end() == self.end_offset:
            return None
        # if ord(self.str_text[m_mae.end()]) == 0x00e8e2:
        #     d.dprint(self.str_text[self.offset:self.offset+10])
        #     d.dprint(format(ord(self.str_text[m_mae.end()]), '#08x'))
        #     d.dprint(format(ord(self.str_text[m_mae.end()+1]), '#08x'))
        # if ord(self.str_text[m_mae.end()]) == 0x00e8e3:
        #     d.dprint(self.str_text[self.offset:self.offset+10])
        #     d.dprint(format(ord(self.str_text[m_mae.end()]), '#08x'))
        #     d.dprint(format(ord(self.str_text[m_mae.end()+1]), '#08x'))
        str_ = ''
        if ord(self.str_text[m_mae.end()]) == 0x00e8e2:
            # d.dprint(format(ord(self.str_text[m_mae.end()]), '#08x'))
            if (0x00e7e8 <= ord(self.str_text[m_mae.end()+1])) \
                    and (ord(self.str_text[m_mae.end()+1]) <= 0x00e7ec):
                m_ushiro = self.p_kakko_ushiro.match(self.str_text,
                        m_mae.end()+2)
                if not m_ushiro:
                    return None
                str_ = m_mae.group(0) + \
                        chr(ord('⑾')+ord(self.str_text[m_mae.end()+1])-0x00e7e8) \
                        + m_ushiro.group(0)
            elif ord(self.str_text[self.offset+1]) == 0x00e7dd:
                m_ushiro = self.p_kakko_ushiro.match(self.str_text,
                        m_mae.end()+2)
                if not m_ushiro:
                    return None
                str_ = m_mae.group(0) + \
                        '⑽' \
                        + m_ushiro.group(0)
            elif (0x00e7d9 <= ord(self.str_text[self.offset+1])) \
                    and (ord(self.str_text[self.offset+1]) <= 0x00e7dc):
                m_ushiro = self.p_kakko_ushiro.match(self.str_text,
                        m_mae.end()+2)
                if not m_ushiro:
                    return m_ushiro
                str_ = m_mae.group(0) + \
                        chr(ord('⒃')+ord(self.str_text[m_mae.end()+1])-0x00e7d9) \
                        + m_ushiro.group(0)
            else:
                return None
        elif ord(self.str_text[m_mae.end()]) == 0x00e8e3:
            if (0x00e7e8 <= ord(self.str_text[m_mae.end()+1])) \
                    and (ord(self.str_text[m_mae.end()+1]) <= 0x00e7ec):
                m_ushiro = self.p_kakko_ushiro.match(self.str_text,
                        m_mae.end()+2)
                if not m_ushiro:
                    return None
                str_ = m_mae.group(0) + '(2' +\
                        chr(ord('1')+ord(self.str_text[m_mae.end()+1])-0x00e7e8) \
                        + ')' +m_ushiro.group(0)
            elif ord(self.str_text[self.offset+1]) == 0xe7dd:
                m_ushiro = self.p_kakko_ushiro.match(self.str_text,
                        m_mae.end()+2)
                if not m_ushiro:
                    return None
                str_ = m_mae.group(0) + \
                        '⒇' \
                        + m_ushiro.group(0)
            elif (0x00e7d9 <= ord(self.str_text[self.offset+1])) \
                    and (ord(self.str_text[self.offset+1]) <= 0x00e7dc):
                m_ushiro = self.p_kakko_ushiro.match(self.str_text,
                        m_mae.end()+2)
                if not m_ushiro:
                    return None
                str_ = m_mae.group(0) + '(2' +\
                        chr(ord('6')+ord(self.str_text[m_mae.end()+1])-0x00e7d9) \
                        + ')' + m_ushiro.group(0)
            else:
                return None
        elif ord(self.str_text[m_mae.end()]) == 0x1a:
            # H21 2009 (21) 農業経営基盤強化準備金制度の改正
            m_ushiro = self.p_kakko_ushiro.match(self.str_text,
                    m_mae.end()+1)
            if not m_ushiro:
                return None
            str_ = m_mae.group(0) + '(21)' +\
                    m_ushiro.group(0)
        else:
            return None
        m = self.p_kakko.match(str_)
        # d.dprint_method_end()
        return (m, m_ushiro.end())


    def create_file_list(self, files, m_dir):
        self.dict_file = {}
        if int(self.str_seireki) != 2005:
            for file in files:
                if int(self.str_seireki) == 2006:
                    num_min = 61
                    num_max = 726
                else:
                    m_filename = self.p_filename.match(file)
                    if m_filename == None:
                        continue
                    num_min = int(m_filename.group(1))
                    num_max = int(m_filename.group(2))
                for i in range(num_min, num_max+1):
                    file_name = os.path.basename(file)
                    str_link = r'https://www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/PDF/' + file_name
                    if int(self.str_seireki) == 2006:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/9551815/www.mof.go.jp/tax_policy/tax_reform/outline/fy2006/' \
                            + file_name
                    if int(self.str_seireki) <= 2015:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/9551815/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/pdf/' + file_name
                    if 2010 <= int(self.str_seireki) <= 2011:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/9551815/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/PDF/' + file_name
                    if 2016 <= int(self.str_seireki) <= 2019:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/11344177/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/pdf/' + file_name
                    if 2020 <= int(self.str_seireki) <= 2020:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/11551246/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/pdf/' + file_name
                    if 2021 <= int(self.str_seireki) <= 2021:
                        str_link = r'https://warp.da.ndl.go.jp/info:ndljp/pid/11719722/www.mof.go.jp/tax_policy/tax_reform/outline/fy' + \
                            self.str_seireki + r'/explanation/' + file_name
                    if 2022 <= int(self.str_seireki) <= 2022:
                        str_link = r'https://warp.ndl.go.jp/web/20230901113105/www.mof.go.jp/tax_policy/tax_reform/outline/fy' +\
                            self.str_seireki + r'/explanation/PDF/' + file_name
                    str_file = './' + self.str_wareki+'_' + \
                            self.str_seireki + '/' + file_name
                    self.dict_file[i] = (file, num_min, num_max,
                            str_link, str_file, file_name)
        else:
            # H17 2005
            for file, num_min, num_max in zip(files,
                    [ 49, 64, 92, 111, 137, 160, 179, 205, 243, 286, 321, 338, 353, 371, 383, 390, 430, 462],
                    [ 63, 91, 110, 136, 159, 178, 204, 242, 285, 320, 337, 352, 370, 382, 389, 429, 461, 480]):
                for i in range(num_min, num_max+1):
                    file_name = os.path.basename(file)
                    str_link = r'https://warp.ndl.go.jp/info:ndljp/pid/1022127/www.mof.go.jp/finance/f1708betu/' + \
                            file_name
                    str_file = './' + self.str_wareki+'_' + \
                            self.str_seireki + '/' + file_name
                    self.dict_file[i] = (file, num_min, num_max,
                            str_link, str_file, file_name)
        
    def make_pdf_to_index(self):
        if not os.path.exists(self.base_folder):
            print("フォルダ {} がありません".self.base_folder)
            return
        
        dirs = glob.glob(self.base_folder + "/*/")
        
        wb = openpyxl.Workbook()
        self.ws = wb.active
        self.ws.append(('和暦', '西暦', '税法等',
                '第', '項目',
                '漢数字', '項目',
                '数字', '項目',
                'カッコ', '項目',
                '丸数字', '項目',
                '全項目',
                '頁', 'リンク'
                ))
        self.ws_row = 2
        
        for str_dir in dirs:
            list_dir = str_dir.split('\\')
            m_dir = self.p_dir.fullmatch(list_dir[-2])
            if m_dir == None:
                continue
            self.str_wareki = m_dir.group(1)
            self.str_seireki = m_dir.group(2)
            # test============================================
            # if self.str_seireki != '2005':
            #     continue
            # test============================================
            if int(self.str_seireki) != 2005:
                # pxxxx-pyyyy.pdf, Pxxx-Pyyy.pdf
                files = glob.glob(str_dir + '*-*.pdf')
                files2 = glob.glob(str_dir + '*_*.pdf')
                files.extend(files2)
                files.sort()
    
                self.create_file_list(files, m_dir)
                            
                out_text = self.base_folder + '/' + list_dir[-2] + '.txt'
                if int(self.str_seireki) != 2006:
                    self.make_each_pdf_to_text(files[1], out_text)
                else:
                    # self.make_each_pdf_to_text("p011-p061.pdf", out_text)
                    self.make_each_pdf_to_text(files[0], out_text)
                self.make_each_text_to_index(out_text)
            else:
                files = [ str_dir + "2.pdf", str_dir + "3.pdf",
                         str_dir + "4.pdf", str_dir + "5.pdf", str_dir + "6.pdf",
                         str_dir + "7.pdf", str_dir + "8.pdf", str_dir + "9.pdf",
                         str_dir + "10.pdf", str_dir + "11.pdf",
                         str_dir + "12.pdf", str_dir + "13.pdf",
                         str_dir + "14.pdf", str_dir + "15.pdf", str_dir + "16.pdf",
                         str_dir + "17.pdf", str_dir + "18.pdf", str_dir + "19.pdf" ]
                self.create_file_list(files, m_dir)
                out_text = self.base_folder + '/' + list_dir[-2] + '.txt'
                self.make_each_pdf_to_text(str_dir + "1-2.pdf", out_text)
                self.make_each_text_to_index(out_text)

        str_filter = "A1:N" + str(self.ws_row - 1)
        self.ws.auto_filter.ref = str_filter
        self.ws.freeze_panes = 'D2'
        self.ws.column_dimensions['A'].width = 4.5
        self.ws.column_dimensions['B'].width = 4.5
        self.ws.column_dimensions['C'].width = 40
        self.ws.column_dimensions['D'].width = 5
        self.ws.column_dimensions['E'].width = 25
        self.ws.column_dimensions['F'].width = 3.25
        self.ws.column_dimensions['G'].width = 65
        self.ws.column_dimensions['H'].width = 3.25
        self.ws.column_dimensions['I'].width = 65
        self.ws.column_dimensions['J'].width = 3.25
        self.ws.column_dimensions['K'].width = 65
        self.ws.column_dimensions['L'].width = 3.25
        self.ws.column_dimensions['M'].width = 65
        self.ws.column_dimensions['N'].width = 65
        self.ws.column_dimensions['O'].width = 4.5
        self.ws.column_dimensions['P'].width = 25
        
        out_excel =  self.base_folder + '/' + '税制改正の解説' + '.xlsx'
        wb.save(out_excel)

     
    def make_each_text_to_index(self, in_text):
        # d.dprint_method_start()
        # d.dprint(self.str_seireki)
        f_text = open(in_text, "r", encoding="utf_8")
        self.str_text = f_text.read()
        f_text.close()
        
        self.end_offset = len(self.str_text)
        if int(self.str_seireki) == 2006:
            self.p_suji = self.p_suji_2006
        elif int(self.str_seireki) == 2005:
            self.p_suji = self.p_suji_2006
        else:
            self.p_suji = self.p_suji_2007e
        for m_low in self.p_low.finditer(self.str_text):
            # 詳解、ページなどを目印に処理を開始
            # d.dprint(m_low.groups())
            # str_low : 所得税関係のその他の改正
            self.str_low = m_low.group(4)
            if int(self.str_seireki) == 2005:
                # 文字列内の空白を削除
                self.str_low = self.str_low.replace(' ', '')
            
            self.before_page = ''
            self.num_page = ''
            
            self.offset = m_low.end()

            self.dai = ('', '')
            self.kansuji = ('', '')
            self.suji = ('', '')
            self.kakko = ('', '')
            self.maru = ('', '')

            while self.offset != self.end_offset:
                # d.dprint("*"+self.str_text[self.offset:self.offset+50])
                # if self.str_text[self.offset:self.offset+50] == '':
                #     d.dprint(self.str_seireki)
                #     d.dprint(self.str_low)
                #     d.dprint(self.str_text[self.offset:self.offset+20])
                #     exit(-1)
                m_dai = self.p_dai.match(self.str_text, self.offset)
                if m_dai:
                    self.make_dai(m_dai)
                else:
                    m_kansuji = self.p_kansuji.match(
                            self.str_text, self.offset)
                    if m_kansuji:
                        self.make_kansuji(m_kansuji)
                    else:
                        m_suji = self.p_suji.match(
                                self.str_text, self.offset)
                        if m_suji:
                            self.make_suji(m_suji)
                        else:
                            m_kakko = self.p_kakko.match(
                                    self.str_text, self.offset)
                            if m_kakko:
                                self.make_kakko(m_kakko)
                            else:
                                ret = self.match_kakko()
                                if ret:
                                    (m_kakko, new_offset) = ret
                                    self.make_kakko_21(m_kakko, new_offset)
                                else:
                                    m_maru = self.p_maru.match(
                                        self.str_text, self.offset)
                                    if m_maru:
                                        self.make_maru(m_maru)
                                    else:
                                        m_page = self.p_page.match(
                                                self.str_text, self.offset)
                                        self.proc_page_etc(m_page)
                                        if not m_page:
                                            m_through = self.p_through.match(
                                                    self.str_text, self.offset)
                                            if m_through:
                                                self.offset = m_through.end()
                                                self.before_page = ''

                m_low = self.p_low.match(self.str_text, self.offset)
                if m_low:
                    break

    def make_fukusu_gyou(self):
        list_koumoku = []
        while True:
            m_low = self.p_low.match(self.str_text, self.offset)
            if m_low:
                break
            m_dai = self.p_dai.match(self.str_text, self.offset)
            if m_dai:
                break
            m_kansuji = self.p_kansuji.match(
                    self.str_text, self.offset)
            if m_kansuji:
                break
            m_suji = self.p_suji.match(
                    self.str_text, self.offset)
            if m_suji:
                break
            m_kakko = self.p_kakko.match(
                    self.str_text, self.offset)
            if m_kakko:
                break
            m_kakko = self.match_kakko()
            if m_kakko:
                break
            m_maru = self.p_maru.match(
                    self.str_text, self.offset)
            if m_maru:
                break
            m_page = self.p_page.match(self.str_text, self.offset)
            if m_page:
                break
            m_through = self.p_through.match(self.str_text, self.offset)
            if not m_through:
                # d.dprint(self.str_seireki)
                # d.dprint(self.str_low)
                # d.dprint(self.str_text[self.offset:self.offset+20])
                break
            if m_through.group(0)[0] == '　':
                break
            # list_koumoku.append('\n')
            list_koumoku.append(m_through.group(0)[:-1])
            self.offset = m_through.end()
        return list_koumoku
        
    def make_dai(self, m):
        self.offset = m.end()
        self.kansuji = (' ', '')
        self.suji = (' ', '')
        self.kakko = (' ', '')
        self.maru = (' ', '')
        self.make_page()
        
        list_koumoku = self.make_fukusu_gyou()
        str_koumoku = m.group(2) + ''.join(list_koumoku)
        self.dai = (m.group(1), str_koumoku)
        
        data_tuple = (self.str_wareki, self.str_seireki,
            self.str_low,
            self.dai[0], self.dai[1],
            self.kansuji[0], self.kansuji[1],
            self.suji[0], self.suji[1],
            self.kakko[0], self.kakko[1],
            self.maru[0], self.maru[1],
            self.num_page)
        if self.num_page != '':
            self.write_excel(data_tuple)
        
    def make_kansuji(self, m):
        self.offset = m.end()
        self.suji = (' ', '')
        self.kakko = (' ', '')
        self.maru = (' ', '')
        self.make_page()
        
        list_koumoku = self.make_fukusu_gyou()
        str_koumoku = m.group(2) + ''.join(list_koumoku)
        self.kansuji = (m.group(1), str_koumoku)
 
        data_tuple = (self.str_wareki, self.str_seireki,
                self.str_low,
                self.dai[0], self.dai[1],
                self.kansuji[0], self.kansuji[1],
                self.suji[0], self.suji[1],
                self.kakko[0], self.kakko[1],
                self.maru[0], self.maru[1],
                self.num_page)
        if self.num_page != '':
            self.write_excel(data_tuple)
        
    def make_suji(self, m):
        self.offset = m.end()
        # 全角から半角に
        self.suji = (m.group(1).translate(self.trans_zenhan), m.group(2))
        self.kakko = (' ', '')
        self.maru = (' ', '')
        self.make_page()
        
        list_koumoku = self.make_fukusu_gyou()
        str_koumoku = m.group(2) + ''.join(list_koumoku)
        # 全角から半角に
        self.suji = (m.group(1).translate(self.trans_zenhan), str_koumoku)

        data_tuple = (self.str_wareki, self.str_seireki,
                self.str_low,
                self.dai[0], self.dai[1],
                self.kansuji[0], self.kansuji[1],
                self.suji[0], self.suji[1],
                self.kakko[0], self.kakko[1],
                self.maru[0], self.maru[1],
                self.num_page)
        if self.num_page != '':
            self.write_excel(data_tuple)
    
    def replace_kakko(self, src):
        dst = src \
                .replace('⑴', '(1)') \
                .replace('⑵', '(2)') \
                .replace('⑶', '(3)') \
                .replace('⑷', '(4)') \
                .replace('⑸', '(5)') \
                .replace('⑹', '(6)') \
                .replace('⑺', '(7)') \
                .replace('⑻', '(8)') \
                .replace('⑼', '(9)') \
                .replace('⑽', '(10)') \
                .replace('⑾', '(11)') \
                .replace('⑿', '(12)') \
                .replace('⒀', '(13)') \
                .replace('⒁', '(14)') \
                .replace('⒂', '(15)') \
                .replace('⒃', '(16)') \
                .replace('⒄', '(17)') \
                .replace('⒅', '(18)') \
                .replace('⒆', '(19)') \
                .replace('⒇', '(20)')
        return dst
    
    def make_kakko(self, m):
        self.offset = m.end()
        self.maru = (' ', '')
        self.make_page()
        
        str_kakko = self.replace_kakko(m.group(1))
        list_koumoku = self.make_fukusu_gyou()
        str_koumoku = m.group(2) + ''.join(list_koumoku)
        self.kakko = (str_kakko, str_koumoku)

        data_tuple = (self.str_wareki, self.str_seireki,
                self.str_low,
                self.dai[0], self.dai[1],
                self.kansuji[0], self.kansuji[1],
                self.suji[0], self.suji[1],
                self.kakko[0], self.kakko[1],
                self.maru[0], self.maru[1],
                self.num_page)
        if self.num_page != '':
            self.write_excel(data_tuple)

    def make_kakko_21(self, m, new_offset):
        self.offset = new_offset
        self.maru = ('', '')
        self.make_page()
        
        list_koumoku = self.make_fukusu_gyou()
        str_koumoku = m.group(2) + ''.join(list_koumoku)
        self.kakko = (m.group(1), str_koumoku)

        data_tuple = (self.str_wareki, self.str_seireki,
                self.str_low,
                self.dai[0], self.dai[1],
                self.kansuji[0], self.kansuji[1],
                self.suji[0], self.suji[1],
                self.kakko[0], self.kakko[1],
                self.maru[0], self.maru[1],
                self.num_page)
        if self.num_page != '':
            self.write_excel(data_tuple)

    def make_maru(self, m):
        self.offset = m.end()
        self.make_page()
        
        list_koumoku = self.make_fukusu_gyou()
        str_koumoku = m.group(2) + ''.join(list_koumoku)
        self.maru = (m.group(1), str_koumoku)

        data_tuple = (self.str_wareki, self.str_seireki,
                self.str_low,
                self.dai[0], self.dai[1],
                self.kansuji[0], self.kansuji[1],
                self.suji[0], self.suji[1],
                self.kakko[0], self.kakko[1],
                self.maru[0], self.maru[1],
                self.num_page)
        if self.num_page != '':
            self.write_excel(data_tuple)
    
    def make_page(self):
        m_page = self.p_page.match(self.str_text, self.offset)
        if not m_page:
            if self.num_page == '':
                self.num_page = self.before_page
        else:
            self.offset = m_page.end()
            self.num_page = m_page.group(1).translate(self.trans_zenhan)
            self.before_page = self.num_page
 
    
    def proc_page_etc(self, m):
        if m != None:
            self.offset = m.end()
            self.num_page = m.group(1).translate(self.trans_zenhan)
            self.before_page = self.num_page
        else:             
            self.num_page = self.before_page

    
    def write_excel(self, data_tuple):
        num_page = data_tuple[11+2]
        str_full = data_tuple[4] + \
                data_tuple[6] + \
                data_tuple[8] + \
                data_tuple[10] + \
                data_tuple[12]
        data_list = list(data_tuple)
        data_list.insert(13, str_full)
        data_tuple = tuple(data_list)
        try:
            self.ws.append(data_tuple)
        except:
            index = data_tuple[8+2].find('\x1a')
            if index == -1:
                print("write_excel")
                for c in data_tuple[8+2]:
                    print(hex(ord(c)))
                print(data_tuple[8+2])
                print(data_tuple)
                print("write_excel")
                # exit(99)
            data_list = list(data_tuple)
            data_list[8+2] = data_list[8+2][:index]
            data_tuple = tuple(data_list)
        if int(self.str_seireki) != 2006:
            try:
                int_page = int(num_page) - int(self.dict_file[int(num_page)][1]) + 1
                link_data = self.dict_file[int(num_page)][3] + '#page=' + str(int_page)
                self.ws.cell(column=13+2+1, row=self.ws_row).value = \
                        self.dict_file[int(num_page)][4][2:]
                self.ws.cell(column=13+2+1, row=self.ws_row).hyperlink = link_data
                str_foxid = self.dict_file[int(num_page)][4][2:]
            except Exception as _e:
                int_page = int(num_page) + 1
                self.ws.cell(column=13+2+1, row=self.ws_row).value = \
                        "リンク設定不能"
                str_foxid = "error"
        else:
            int_page = int(num_page) + 1
            link_data = "https://warp.da.ndl.go.jp/info:ndljp/pid/9551815/www.mof.go.jp/tax_policy/tax_reform/outline/fy2006/f1808betu.pdf" \
                    + '#page=' + str(int_page)
            self.ws.cell(column=13+2+1, row=self.ws_row).value = \
                    "H18_2006/f1808betu.pdf"
            self.ws.cell(column=13+2+1, row=self.ws_row).hyperlink = link_data
            str_foxid = "H18_2006/f1808betu.pdf"
        
        # foxit reader用は、下記の２行を有効にする
        self.ws.cell(column=14+2+1, row=self.ws_row).value = str_foxid
        self.ws.cell(column=15+2+1, row=self.ws_row).value = int_page
        # Excelファイルに下記のマクロを設定する
        #  ctlr+lに設定
        '''
        Sub linkpdfpage()
            Worksheets("Sheet").Activate
            Text = "C:\Program Files (x86)\Foxit Software\Foxit PDF Reader\FoxitPDFReader.exe /A page=" + Mid(Cells(Selection(1).Row, 18).Value, 1) + " " + ActiveWorkbook.Path + "/" + Cells(Selection(1).Row, 17).Value
            Shell (Text)
        End Sub
        '''

        self.ws_row += 1

    def make_each_pdf_to_text(self, in_file, out_file):
        # d.dprint_method_start()
        laparams = LAParams()               # パラメータインスタンス
        laparams.boxes_flow = None          # -1.0（水平位置のみが重要）から+1.0（垂直位置のみが重要）default 0.5
        laparams.word_margin = 0.2          # default 0.1
        laparams.char_margin = 2.0          # default 2.0
        laparams.line_margin = 0            # default 0.5
        
        with open(out_file, "w", encoding="utf-8") as f:
            self.text_ = ""
            
            for page_layout in extract_pages(in_file, maxpages=0, laparams=laparams):    # ファイルにwithしている
                for element in sorted(self.flatten_lttext(page_layout, LTTextBox), key=lambda x: (-x.y1, x.x0)):
                    if element.y1 < self.footer:
                        continue  # フッター位置の文字は抽出しない
                    if element.y0 > self.header:
                        continue  # ヘッダー位置の文字は抽出しない
                    _text =element.get_text()
                    if '(cid:' in _text:
                        _text = re.sub(r"\(cid:([0-9]*)\)",
                                lambda m: chr(ord('⑴')+int(m.group(1))-2), _text) 
                    _text = _text.replace(chr(0x18), '9')
                    _text = _text.replace(chr(0x07), '')
                    
                    # 太字のページ数
                    if self.str_seireki == "2009" or self.str_seireki =="2010":
                        if re.fullmatch(r'[0-9]*(\r\n|\n|\r)', _text):
                            _text = _text[int((len(_text)-1)/2):]
                    if self.str_seireki == "2006":
                        _text = _text.translate(self.trans_2006_kakko)
                    if self.str_seireki == "2007":
                        _text = _text.translate(self.trans_2007_kakko)
                        if "i" in _text:
                            _text = _text.replace("i1", "⑾")
                            _text = _text.replace("i2", "⑿")
                            _text = _text.replace("i3", "⒀")
                            _text = _text.replace("i4", "⒁")
                            _text = _text.replace("i5", "⒂")
                            _text = _text.replace("i6", "⒃")
                            _text = _text.replace("i7", "⒄")
                            _text = _text.replace("i8", "⒅")
                            _text = _text.replace("i9", "⒆")
                    self.text_ += _text
                # d.dprint(self.text_)
                self.write2text(f)

    def flatten(self, l):
        """
        ツリー状になっているイテレータをフラットに返すイテレータ
        """
        for el in l:
            if isinstance(el, collections.abc.Iterable) and not isinstance(el, (str, bytes)):
                yield from self.flatten(el)
            else:
                yield el

    def flatten_lttext(self, l, _type):
        """
        ツリー状になっているイテレータをフラットに返すイテレータ
        返る要素の型を指定
        pdfminerのextract_pagesで使用するのを想定
        要素の型が引数で指定した型を継承したもののみを返す

        Args:
            l:      pdfminerのextract_pages()の戻り値
            _type:  戻したい値の型
        """
        for el in l:
            if isinstance(el, (_type)):
                yield el
            else:
                if isinstance(el, collections.abc.Iterable) and not isinstance(el, (str, bytes)):
                    yield from self.flatten_lttext(el, _type)
                else:
                    continue

    def write2text(self, f):
        """
        Args:
            f:      書き込みファイル
        """
        f.write(self.text_)
        self.text_ = ""

    def convert_pdf_to_text(self):
        """
        PDFファイルをテキストに変換
        PDFは2段に段組みされたものも含む
        """

        laparams = LAParams()               # パラメータインスタンス
        laparams.boxes_flow = None          # -1.0（水平位置のみが重要）から+1.0（垂直位置のみが重要）default 0.5
        laparams.word_margin = 0.2          # default 0.1
        laparams.char_margin = 2.0          # default 2.0
        laparams.line_margin = 0            # default 0.5

        with open(self.output_path, "w", encoding="utf-8") as f:
            self.text_l = ""        # 左側の文字列
            self.text_r = ""        # 右側の文字列
            
            print("Analyzing from {} page to {} page(0:to last)".format(self.start_page, self.last_page))
            
            for page_layout in extract_pages(self.input_path, maxpages=0, laparams=laparams):    # ファイルにwithしている
                if page_layout.pageid < self.start_page: continue                   # 指定開始ページより前は飛ばす
                if self.last_page and self.last_page < page_layout.pageid: break    # 指定終了ページ以降は中断
                if self.border == 0:
                    self.border = int(page_layout.width / 2)
                if page_layout.pageid == self.start_page:
                    print("Check on page #{}".format(page_layout.pageid))
                    print("Page Info width:{}, heght:{}".format(page_layout.width, page_layout.height))
                    print("Calc result border: {}, footer: {}".format(self.border, self.footer))
                for element in sorted(self.flatten_lttext(page_layout, LTTextBox), key=lambda x: (-x.y1, x.x0)):
                    if element.y1 < self.footer: continue  # フッター位置の文字は抽出しない
                    if element.y0 > self.header: continue  # ヘッダー位置の文字は抽出しない
                    _text =element.get_text()
                    # debug
                    # print("y1:{}, y0:{}, x0:{}, x1:{}■{}".
                    #       format(element.y1, element.y0, element.x0, element.x1, _text))

                    if element.x1 < self.border:
                        # 文字列全体が左側
                        self.text_l += _text
                    else:
                        if element.x0 >= self.border:
                            # 文字列全体が右側
                            self.text_r += _text
                        else:
                            # 文字列が境界をまたいでいる場合
                            # 右側に既に文章があれば先に出力する
                            if self.text_r:
                                self.write2text(f)
                            self.text_l += _text
                self.write2text(f)

if __name__ == "__main__":
    make = MakeIndex(sys.argv[1:])
    make.make_pdf_to_index()
